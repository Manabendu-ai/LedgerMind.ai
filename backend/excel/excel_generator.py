from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os


class ExcelGenerator:

    def save(self, workbook_json, filename: str, persist_dir: str="excel_files/")->str:

        os.makedirs(persist_dir, exist_ok=True)
        output_path = os.path.join(persist_dir, f"{filename}.xlsx")
        wb = Workbook()

        wb.remove(wb.active)

        header_fill = PatternFill(
            fill_type="solid",
            start_color="1F4E78",
            end_color="1F4E78"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for sheet in workbook_json.workbook.worksheets:

            ws = wb.create_sheet(sheet.worksheet_name)

            
            for col, header in enumerate(sheet.columns, start=1):

                cell = ws.cell(row=1, column=col)

                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

            for row_no, row in enumerate(sheet.rows, start=2):

                for col_no, value in enumerate(row, start=1):

                    cell = ws.cell(row=row_no, column=col_no)

                    cell.value = value
                    cell.border = border

            for column_cells in ws.columns:

                max_length = max(
                    len(str(cell.value)) if cell.value else 0
                    for cell in column_cells
                )

                ws.column_dimensions[
                    column_cells[0].column_letter
                ].width = min(max_length + 4, 60)

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        wb.save(output_path)

        print(f"[SUCCESS] Excel saved at {output_path}")
        return output_path